import datetime
import math

import openpyxl
import pandas as pd
from pathlib import Path
import re


class DataRow:
    """
    class to set the data row
    """
    def __init__(self, date: datetime, product: str, hub: str, price: float, bool_product_type: bool):
        self.date = date
        self.product = product
        if hub.find('/') != -1:
            self.hub = hub[:hub.find('/')]
            self.hub2 = hub[hub.find('/') + 1:]
        else:
            self.hub = hub
            self.hub2 = '-'
        self.price = price
        self.prices_name = hub.replace('/', '_')
        self.unit = 'MWh'
        self.currency = 'EUR'
        self.source = '42 Financial Services'
        self.price_type = 'bid' if bool_product_type else 'offer'
        self.product_types = {'Daily': r'\A(WD|DA|WE)$',
                              'Month': r'\A(BOM|(FEB|MAR|APRIL)\d{2})$',
                              'Quarter': r'\AQ\d{3}$',
                              'Season': r'\A(SUM|WIN)\d{2}$',
                              'Year': r'\ACAL\d{2}$'}
        self.product_type = self.get_product_type()

    def get_product_type(self):
        price_type = None
        for key in self.product_types.keys():
            if re.search(self.product_types[key], self.product):
                price_type = key
                break
        if price_type is not None:
            return price_type
        else:
            raise ValueError('no mask for product ' + self.product)

    def set_df_row(self):
        df_row = {
            'date': self.date,
            'prices_name': self.prices_name,
            'price': self.price,
            'hub': self.hub,
            'hub2': self.hub2,
            'unit': self.unit,
            'currency': self.currency,
            'price_type': self.price_type,
            'products': self.product,
            'source': self.source,
            'product_type': self.product_type
        }
        return df_row


class GASSheet:
    """
    class to set and read the xlsx sheet
    """
    def __init__(self, sheet_):
        self.date = datetime.datetime.strptime(sheet_.title, '%d%b%Y')
        self.df_sheet = pd.DataFrame(columns=['date', 'prices_name', 'price', 'hub', 'hub2', 'unit', 'currency', 'price_type',
                                              'products', 'source', 'product_type'])

        def read_sheet():
            def find_first_col():
                for col_ in sheet_.iter_cols(1, sheet_.max_column):
                    for row_ in range(0, sheet_.max_row):
                        value_ = col_[row_].value
                        if value_ is not None:
                            return col_[row_].column

            first_col = find_first_col()

            def find_first_row():
                for row_ in sheet_.iter_rows(1, sheet_.max_row):
                    value_ = row_[first_col - 1].value
                    if value_ is None:
                        continue
                    elif len(value_.strip()) == 0:
                        continue
                    else:
                        return row_[first_col - 1].row

            first_row = find_first_row()
            hub_row = first_row - 2

            for row_ in sheet_.iter_rows(first_row, sheet_.max_row):
                product = row_[first_col - 1].value.strip()
                if product == 'Time swap':
                    break
                else:
                    for col_ in sheet_.iter_cols(2, sheet_.max_column):
                        if col_[hub_row - 1].value is not None:
                            hub = col_[hub_row - 1].value.strip()
                            price_bid = row_[col_[hub_row - 1].column - 1].value
                            price_ask = row_[col_[hub_row - 1].column].value
                            if price_bid is None:
                                continue
                            else:
                                new_data_row = DataRow(self.date, product, hub, price_bid, True)
                                self.df_sheet.loc[len(self.df_sheet.index)] = new_data_row.set_df_row()
                            if price_ask is None:
                                continue
                            else:
                                new_data_row = DataRow(self.date, product, hub, price_ask, False)
                                self.df_sheet.loc[len(self.df_sheet.index)] = new_data_row.set_df_row()

        read_sheet()


class GASParser:
    """
    class to read the xlsx file
    """
    def __init__(self, file_name):
        self.file_name = file_name
        self.xlsx_file = Path('', self.file_name)
        self.df = pd.DataFrame(columns=['date', 'prices_name', 'price', 'hub', 'hub2', 'unit', 'currency', 'price_type',
                                        'products', 'source', 'product_type'])

        self.get_sheets_from_file()
        #replacing hubs
        self.df = self.df.replace({'Czech Virtual Point': 'Czech VTP',
                                   'THE': 'THE VTP', 'CEGH': 'Austria VTP',
                                   'SK VTP': 'Slovak VTP', 'VTP': 'Austria VTP'})
        #replacing instruments
        self.df = self.df.replace({'APRIL23':'APR23','WE':'W/END', 'CAL24':'2024', 'CAL25':'2025', 'CAL26':'2026','WIN23':'Win 2023/2024',
                                   'SUM23':'Sum 2023', 'SUM24':'Sum 2024', 'WIN24':'Win 2024/2025',
                                   'Q122':'Q1/22', 'Q222':'Q2/22', 'Q322':'Q3/22', 'Q422':'Q4/22', 
                                   'Q123':'Q1/23', 'Q223':'Q2/23', 'Q323':'Q3/23', 'Q423':'Q4/23',
                                   'Q124':'Q1/24', 'Q224':'Q2/24', 'Q324':'Q3/24', 'Q424':'Q4/24',
                                   'Tue': 'Tuesday', 'Wed': 'Wednesday', 'Thu': 'Thursday',
                                   'Fri': 'Friday', 'Sat': 'Saturday', 'Sun': 'Sunday'})

    def get_sheets_from_file(self):
        wb = openpyxl.load_workbook(self.xlsx_file, data_only=True)
        for sheet_ in wb.worksheets:
            new_GAS_sheet = GASSheet(sheet_)
            self.df = pd.concat([self.df, new_GAS_sheet.df_sheet], sort=False, axis=0)
            self.df.reset_index(drop=True, inplace=True)

    def concat_two_df(self, df2):
        new_df = pd.concat([self.df, df2], sort=False, axis=0)
        new_df.reset_index(drop=True, inplace=True)
        return new_df

    def write_df_to_csv(self):
        self.df.to_csv(self.xlsx_file.stem + '.csv')


if __name__ == '__main__':
    file_name = r'ClosingDayPricesGAS2023.xlsx'
    new_parser = GASParser(file_name)
    new_parser.df.to_csv(r'ClosingDayPricesGAS2023.csv')



