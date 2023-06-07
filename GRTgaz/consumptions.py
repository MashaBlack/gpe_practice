import requests
import openpyxl
from pathlib import Path
from datetime import datetime, timedelta
from copy import copy


class ConsumptionsParser:
    @staticmethod
    def get_data_from_grtgaz(file_name: str, params: dict):
        """
        function to get .xlsx data from GRTgaz
        """
        url = 'https://www.smart.grtgaz.com/api/v1/en/consommation/export/Zone.xls'
        r = requests.get(url, params=params)
        with open(file_name + '.xlsx', 'wb') as f:
            f.write(r.content)

    @staticmethod
    def get_current_data(file_name):
        """
        get data for the past two days
        """
        file_name = file_name + '_' + str(datetime.now().date())
        params = {
            'startDate': str((datetime.now() - timedelta(days=2)).date()),
            'endDate': str(datetime.now().date()),
            'range': 'daily'
        }
        ConsumptionsParser.get_data_from_grtgaz(file_name, params)
        xls = XLSXParser(file_name)
        xls.save()

    @staticmethod
    def get_historical_data(file_name):
        """
        function to get historical data from 01.04.2015 till now into .xlsx file
        """
        start_year = 2015
        end_year = datetime.now().year

        file_name = file_name + '_' + str(start_year) + '–' + str(end_year)

        def copy_style(copy_cell, new_cell):
            """
            function to copy cell style
            """
            new_cell.font = copy(copy_cell.font)
            new_cell.border = copy(copy_cell.border)
            new_cell.fill = copy(copy_cell.fill)
            new_cell.number_format = copy(copy_cell.number_format)
            new_cell.protection = copy(copy_cell.protection)
            new_cell.alignment = copy(copy_cell.alignment)

        def collect_all_data(xls: XLSXParser, temp_xls: XLSXParser):
            """
            collects data from temp xls into xls with all data
            """
            start_row = 4
            max_row = xls.sheet.max_row
            for iter_row in range(start_row, temp_xls.sheet.max_row + 1):
                for iter_col in range(1, xls.sheet.max_column + 1):
                    new_cell = xls.sheet.cell(row=max_row + iter_row - start_row + 1, column=iter_col)
                    copy_cell = temp_xls.sheet.cell(row=iter_row, column=iter_col)
                    new_cell.value = copy_cell.value
                    copy_style(copy_cell, new_cell)

        openpyxl.Workbook().save(file_name + '.xlsx')  # save empty xls
        xls = XLSXParser(file_name)

        for year in range(start_year, end_year + 1):
            params = {
                'startDate': str(year) + '-01-01',
                'endDate': str(year) + '-12-31',
                'range': 'daily'
            }
            if year == start_year:
                params['startDate'] = str(year) + '-04-01'
            else:
                file_name = 'temp'  # all files except the first are temp

            if year == end_year:
                params['endDate'] = str(datetime.now().date())

            ConsumptionsParser.get_data_from_grtgaz(file_name, params)

            if year == start_year:
                xls = XLSXParser(file_name)
            else:
                temp_xls = XLSXParser(file_name)
                collect_all_data(xls, temp_xls)

        Path('temp.xlsx').unlink()  # delete temp file
        xls.save()


class XLSXParser:
    """
    class to read the xlsx file
    """
    def __init__(self, file_name):
        self.file_name = file_name
        self.xlsx_file = Path('', self.file_name + '.xlsx')
        self.wb = openpyxl.load_workbook(self.xlsx_file, data_only=True)
        self.sheet = self.wb.active
        self.get_certain_xls_view()

    def change_sheet_view(self):
        """
        function to change sheet view as in the example
        (delete extra columns and rename the others)
        """
        self.sheet.delete_cols(2, 1)
        self.sheet.delete_cols(6, self.sheet.max_column - 5)
        new_col_names = ['FR_Industrial demand', 'FR_LDZ demand', 'FR_Power demand', 'FR_Other demand']
        for i in range(len(new_col_names)):
            self.sheet.cell(row=3, column=i + 2).value = new_col_names[i]

    def get_certain_xls_view(self):
        """
        function to delete extra sheets and to change active sheet
        """
        for sheet in self.wb.worksheets:
            if sheet is not self.sheet:
                self.wb.remove(sheet)
            else:
                self.change_sheet_view()

    def save(self):
        """
        functon to save .xlsx file
        """
        self.wb.save(self.file_name + '.xlsx')


if __name__ == '__main__':
    parser = ConsumptionsParser()
    parser.get_historical_data('all_GRTgaz_consumptions')
    parser.get_current_data('GRTgaz_consumptions')
